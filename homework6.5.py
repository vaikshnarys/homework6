import csv
from openpyxl import Workbook

list_data = []
with open("first.csv","r") as new_file:
    reader = csv.reader(new_file)
    headers = next(reader)
    for row in reader:
        list_data.append(row[:2] + row[3:])

wb = Workbook()
ws = wb.active

for row_index,value in enumerate(headers[:3],start = 2):
    ws.cell(row = row_index, column = 1 , value = value)
    ws.cell(row = 4, column = 1, value = headers[3])

for col_index,item in enumerate(range(1,7),start = 2):
    ws.cell(row = 1, column = col_index, value = f"Person{item}" )
for col_index, item in enumerate(list_data, start=2):
    for row_index, value in enumerate(item, start=2):
        ws.cell(row=row_index, column=col_index, value=value)

wb.save("first.xlsx")